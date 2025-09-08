from PyQt5.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton, QComboBox, QFileDialog, QMessageBox, QTextEdit, QSpinBox, QScrollArea, QWidget, QCheckBox, QGroupBox, QVBoxLayout, QToolButton, QFrame, QTextBrowser, QProgressDialog
import openpyxl
from openpyxl import load_workbook
from qgis.core import QgsProject
import traceback  # Import traceback for detailed error information
from PyQt5.QtCore import Qt, QUrl
from PyQt5.QtGui import QDesktopServices
import os
from urllib.parse import unquote  # Import for decoding URL-encoded characters
from qgis.gui import QgsCollapsibleGroupBox  # Import QgsCollapsibleGroupBox

class CustomTextBrowser(QTextBrowser):
    def setSource(self, url: QUrl, type=None):
        """
        Override the default behavior of QTextBrowser to prevent clearing the console.
        """
        if url.isLocalFile():
            QDesktopServices.openUrl(url)  # Open the local file in the default application
        else:
            super().setSource(url, type)

class ValidateProjectReportDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Validate Project Report")
        self.setMinimumWidth(400)

        # Load cached selections from QGIS project variables
        self.project = QgsProject.instance()
        self.cache_prefix = "ValidateProjectReportDialog"

        layout = QVBoxLayout(self)

        # Excel file selection
        excel_file_layout = QHBoxLayout()
        self.excel_file_edit = QLineEdit(self)
        self.excel_file_edit.setText(self.get_cached_value("excel_file", ""))
        browse_button = QPushButton("Browse", self)
        browse_button.clicked.connect(self.browse_excel_file)
        excel_file_layout.addWidget(QLabel("Excel File:", self))
        excel_file_layout.addWidget(self.excel_file_edit)
        excel_file_layout.addWidget(browse_button)
        layout.addLayout(excel_file_layout)

        # Worksheet selection
        self.sheet_combo = QComboBox(self)
        self.sheet_combo.currentIndexChanged.connect(self.populate_headers)
        layout.addWidget(QLabel("Select Worksheet:"))
        layout.addWidget(self.sheet_combo)

        # Header row selection
        header_row_layout = QHBoxLayout()
        self.header_row_spin = QSpinBox(self)
        self.header_row_spin.setMinimum(1)
        self.header_row_spin.setValue(int(self.get_cached_value("header_row", "1")))
        self.header_row_spin.valueChanged.connect(self.populate_headers)
        header_row_layout.addWidget(QLabel("Header Row:"))
        header_row_layout.addWidget(self.header_row_spin)
        layout.addLayout(header_row_layout)

        # Max columns selection
        max_columns_layout = QHBoxLayout()
        self.max_columns_spin = QSpinBox(self)
        self.max_columns_spin.setMinimum(1)
        self.max_columns_spin.setValue(int(self.get_cached_value("max_columns", "20")))
        self.max_columns_spin.valueChanged.connect(self.populate_headers)
        max_columns_layout.addWidget(QLabel("Max Columns:"))
        max_columns_layout.addWidget(self.max_columns_spin)
        layout.addLayout(max_columns_layout)

        # Layer name field selection
        self.layer_name_combo = QComboBox(self)
        layout.addWidget(QLabel("Layer Name Field:"))
        layout.addWidget(self.layer_name_combo)

        # Source path field selection
        self.source_path_combo = QComboBox(self)
        layout.addWidget(QLabel("Source Path Field:"))
        layout.addWidget(self.source_path_combo)

        # Layer Name Descriptor Delimiter
        layer_name_delimiter_layout = QHBoxLayout()
        self.layer_name_delimiter_edit = QLineEdit(self)
        self.layer_name_delimiter_edit.setText(self.get_cached_value("layer_name_delimiter", "_"))  # Default to "_"
        layer_name_delimiter_layout.addWidget(QLabel("Layer Name Descriptor Delimiter:"))
        layer_name_delimiter_layout.addWidget(self.layer_name_delimiter_edit)
        layout.addLayout(layer_name_delimiter_layout)

        # Case-Sensitive Layer Matching
        self.case_sensitive_checkbox = QCheckBox("Case-Sensitive Layer Matching", self)
        self.case_sensitive_checkbox.setChecked(False)  # Default to unchecked
        layout.addWidget(self.case_sensitive_checkbox)

        # Verbose Console option
        self.verbose_console_checkbox = QCheckBox("Verbose Console", self)
        self.verbose_console_checkbox.setChecked(self.get_cached_value("verbose_console", "False") == "True")
        layout.addWidget(self.verbose_console_checkbox)

        # Collapsible panel for filter categories
        self.filter_group_box = QgsCollapsibleGroupBox(self)
        self.filter_group_box.setTitle("Filter Categories")
        self.filter_group_box.setCollapsed(False)  # Default to expanded
        filter_layout = QVBoxLayout(self.filter_group_box)

        # First filter category
        self.use_filter_category_checkbox1 = QCheckBox("Enable First Filter Category", self)
        self.use_filter_category_checkbox1.setChecked(True)
        self.use_filter_category_checkbox1.stateChanged.connect(self.toggle_filter_category1)
        filter_layout.addWidget(self.use_filter_category_checkbox1)

        self.filter_category_combo1 = QComboBox(self)
        self.filter_category_combo1.currentIndexChanged.connect(self.populate_filter_categories1)
        filter_layout.addWidget(QLabel("First Filter Category Field:"))
        filter_layout.addWidget(self.filter_category_combo1)

        self.filter_category_scroll1 = QScrollArea(self)
        self.filter_category_scroll1.setWidgetResizable(True)
        self.filter_category_widget1 = QWidget()
        self.filter_category_layout1 = QVBoxLayout(self.filter_category_widget1)
        self.filter_category_scroll1.setWidget(self.filter_category_widget1)
        filter_layout.addWidget(QLabel("Select First Filter Categories:"))
        filter_layout.addWidget(self.filter_category_scroll1)

        # Second filter category
        self.use_filter_category_checkbox2 = QCheckBox("Enable Second Filter Category", self)
        self.use_filter_category_checkbox2.setChecked(False)
        self.use_filter_category_checkbox2.stateChanged.connect(self.toggle_filter_category2)
        filter_layout.addWidget(self.use_filter_category_checkbox2)

        self.filter_category_combo2 = QComboBox(self)
        self.filter_category_combo2.currentIndexChanged.connect(self.populate_filter_categories2)
        filter_layout.addWidget(QLabel("Second Filter Category Field:"))
        filter_layout.addWidget(self.filter_category_combo2)

        self.filter_category_scroll2 = QScrollArea(self)
        self.filter_category_scroll2.setWidgetResizable(True)
        self.filter_category_widget2 = QWidget()
        self.filter_category_layout2 = QVBoxLayout(self.filter_category_widget2)
        self.filter_category_scroll2.setWidget(self.filter_category_widget2)
        filter_layout.addWidget(QLabel("Select Second Filter Categories:"))
        filter_layout.addWidget(self.filter_category_scroll2)

        layout.addWidget(self.filter_group_box)

        # Duplicate match mode selection
        duplicate_match_layout = QHBoxLayout()
        self.duplicate_match_combo = QComboBox(self)
        self.duplicate_match_combo.addItems(["STOP ON FIRST SOURCE MATCH", "STOP ON FIRST LAYER MATCH"])
        self.duplicate_match_combo.setCurrentText(self.get_cached_value("duplicate_match_mode", "STOP ON FIRST SOURCE MATCH"))
        duplicate_match_layout.addWidget(QLabel("Duplicate Match Mode:", self))
        duplicate_match_layout.addWidget(self.duplicate_match_combo)
        layout.addLayout(duplicate_match_layout)

        # Validation report path selection
        report_path_layout = QHBoxLayout()
        self.report_path_edit = QLineEdit(self)
        self.report_path_edit.setText(self.get_cached_value("report_path", ""))
        browse_report_button = QPushButton("Browse", self)
        browse_report_button.clicked.connect(self.browse_report_path)
        report_path_layout.addWidget(QLabel("Validation Report Path:", self))
        report_path_layout.addWidget(self.report_path_edit)
        report_path_layout.addWidget(browse_report_button)
        layout.addLayout(report_path_layout)

        # Option to generate an HTML report
        self.generate_html_checkbox = QCheckBox("Generate HTML Report", self)
        self.generate_html_checkbox.setChecked(False)  # Default to unchecked
        layout.addWidget(self.generate_html_checkbox)

        # Debug console log (use CustomTextBrowser for clickable links)
        self.log_console = CustomTextBrowser(self)
        self.log_console.setOpenExternalLinks(False)  # Disable automatic handling of external links
        layout.addWidget(QLabel("Console:"))
        layout.addWidget(self.log_console)

        # OK and Cancel buttons
        button_layout = QHBoxLayout()
        ok_button = QPushButton("Create Report", self)
        ok_button.clicked.connect(self.validate_project)
        cancel_button = QPushButton("Exit", self)
        cancel_button.clicked.connect(self.reject)
        button_layout.addWidget(ok_button)
        button_layout.addWidget(cancel_button)
        layout.addLayout(button_layout)

        # Show the dialog immediately and restore cached selections in the background
        self.show()
        self.restore_cached_selections_with_progress()

    def toggle_filter_category1(self, state):
        """
        Enable or disable the first filter category dropdown and scrollable box based on the checkbox state.
        """
        enabled = state == Qt.Checked
        self.filter_category_combo1.setEnabled(enabled)
        self.filter_category_scroll1.setEnabled(enabled)
        self.log_message(f"First filter category {'enabled' if enabled else 'disabled'}.", debug=True)

    def toggle_filter_category2(self, state):
        """
        Enable or disable the second filter category dropdown and scrollable box based on the checkbox state.
        """
        enabled = state == Qt.Checked
        self.filter_category_combo2.setEnabled(enabled)
        self.filter_category_scroll2.setEnabled(enabled)
        self.log_message(f"Second filter category {'enabled' if enabled else 'disabled'}.", debug=True)

    def restore_cached_selections_with_progress(self):
        """
        Restore cached selections and show a progress bar for the process.
        """
        tasks = [
            ("Loading Excel file", self.load_cached_excel_file),
            ("Loading worksheet", self.load_cached_worksheet),
            ("Populating headers", self.populate_cached_headers),
            ("Restoring filter categories", self.restore_cached_filter_categories),
            ("Restoring other settings", self.restore_other_cached_settings),
        ]

        progress_dialog = QProgressDialog("Restoring cached selections...", "Cancel", 0, len(tasks), self)
        progress_dialog.setWindowTitle("Loading")
        progress_dialog.setWindowModality(Qt.WindowModal)
        progress_dialog.setMinimumDuration(0)

        for i, (description, task) in enumerate(tasks):
            if progress_dialog.wasCanceled():
                break
            progress_dialog.setLabelText(description)
            progress_dialog.setValue(i)
            task()  # Execute the task
        progress_dialog.setValue(len(tasks))

    def load_cached_excel_file(self):
        """
        Load the cached Excel file path and populate the file edit box.
        """
        excel_file = self.get_cached_value("excel_file", "")
        if excel_file:
            self.excel_file_edit.setText(excel_file)
            self.load_sheets(excel_file)

    def load_cached_worksheet(self):
        """
        Load the cached worksheet and set it in the combo box.
        """
        cached_sheet = self.get_cached_value("sheet", "")
        if cached_sheet and cached_sheet in [self.sheet_combo.itemText(i) for i in range(self.sheet_combo.count())]:
            self.sheet_combo.setCurrentText(cached_sheet)

    def populate_cached_headers(self):
        """
        Populate headers and validate cached fields.
        """
        self.populate_headers()
        cached_layer_name = self.get_cached_value("layer_name_field", "")
        cached_source_path = self.get_cached_value("source_path_field", "")

        if cached_layer_name and cached_layer_name in [self.layer_name_combo.itemText(i) for i in range(self.layer_name_combo.count())]:
            self.layer_name_combo.setCurrentText(cached_layer_name)

        if cached_source_path and cached_source_path in [self.source_path_combo.itemText(i) for i in range(self.source_path_combo.count())]:
            self.source_path_combo.setCurrentText(cached_source_path)

    def restore_cached_filter_categories(self):
        """
        Restore cached filter categories and their selections.
        """
        use_filter_category1 = self.get_cached_value("use_filter_category1", "True") == "True"
        self.use_filter_category_checkbox1.setChecked(use_filter_category1)

        cached_filter_category_field1 = self.get_cached_value("filter_category_field1", "")
        if use_filter_category1 and cached_filter_category_field1 and cached_filter_category_field1 in [self.filter_category_combo1.itemText(i) for i in range(self.filter_category_combo1.count())]:
            self.filter_category_combo1.setCurrentText(cached_filter_category_field1)

            # Populate filter categories and validate cached selections
            self.populate_filter_categories1()
            cached_filter_categories1 = self.get_cached_value("filter_categories1", "").split(",")
            for i in range(self.filter_category_layout1.count()):
                checkbox = self.filter_category_layout1.itemAt(i).widget()
                if isinstance(checkbox, QCheckBox) and checkbox.text() in cached_filter_categories1:
                    checkbox.setChecked(True)

        use_filter_category2 = self.get_cached_value("use_filter_category2", "False") == "True"
        self.use_filter_category_checkbox2.setChecked(use_filter_category2)

        cached_filter_category_field2 = self.get_cached_value("filter_category_field2", "")
        if use_filter_category2 and cached_filter_category_field2 and cached_filter_category_field2 in [self.filter_category_combo2.itemText(i) for i in range(self.filter_category_combo2.count())]:
            self.filter_category_combo2.setCurrentText(cached_filter_category_field2)

            # Populate second filter categories and validate cached selections
            self.populate_filter_categories2()
            cached_filter_categories2 = self.get_cached_value("filter_categories2", "").split(",")
            for i in range(self.filter_category_layout2.count()):
                checkbox = self.filter_category_layout2.itemAt(i).widget()
                if isinstance(checkbox, QCheckBox) and checkbox.text() in cached_filter_categories2:
                    checkbox.setChecked(True)

    def restore_other_cached_settings(self):
        """
        Restore other cached settings such as layer name delimiter, case sensitivity, etc.
        """
        self.layer_name_delimiter_edit.setText(self.get_cached_value("layer_name_delimiter", "_"))
        self.case_sensitive_checkbox.setChecked(self.get_cached_value("case_sensitive_matching", "False") == "True")
        self.generate_html_checkbox.setChecked(self.get_cached_value("generate_html_report", "False") == "True")
        self.verbose_console_checkbox.setChecked(self.get_cached_value("verbose_console", "False") == "True")

        # Set default report path if not restored from settings
        cached_report_path = self.get_cached_value("report_path", "")
        if cached_report_path:
            self.report_path_edit.setText(cached_report_path)
        else:
            qgis_project_path = QgsProject.instance().fileName()
            if qgis_project_path:
                project_dir = os.path.dirname(qgis_project_path)
                project_name = os.path.splitext(os.path.basename(qgis_project_path))[0]
                default_report_path = os.path.join(project_dir, f"{project_name}_ValidationReport.csv")
                self.report_path_edit.setText(default_report_path)
                self.log_message(f"Default report path set to: {default_report_path}", debug=True)
            else:
                self.log_message("No QGIS project loaded. Unable to set default report path.")

    def get_cached_value(self, key, default):
        """
        Retrieve a cached value from QGIS project variables.
        If the value is empty or invalid, return the default value.
        """
        group_name = self.cache_prefix
        value, ok = self.project.readEntry(group_name, key)
        if not ok or not value:  # Check for empty or invalid values
            return default
        return value

    def save_cached_value(self, key, value):
        """
        Save a value to QGIS project variables.
        """
        group_name = self.cache_prefix
        try:
            self.project.writeEntry(group_name, key, str(value))  # Save the value as a string
            self.log_message(f"Saved project variable: {group_name}/{key} = {value}",debug=True)
        except Exception as e:
            self.log_message(f"Failed to save project variable: {group_name}/{key}. Error: {e}")

    def log_message(self, message, debug=False):
        """
        Append a message to the debug console log.
        If `debug` is True, the message will only be logged if "Verbose Console" is enabled.
        """
        if debug and not self.verbose_console_checkbox.isChecked():
            return  # Skip debug messages if "Verbose Console" is disabled
        self.log_console.append(message)
        self.log_console.ensureCursorVisible()  # Ensure the console scrolls to the bottom

    def log_message_link(self, message, link):
        """
        Append a clickable link to the debug console log.
        :param message: The display text for the link.
        :param link: The actual link (file path or URL).
        """
        file_url = QUrl.fromLocalFile(link).toString()
        format_message = message + f" ({link})"
        self.log_console.append(f'<a href="{file_url}">{format_message}</a>')
        self.log_console.ensureCursorVisible()  # Ensure the console scrolls to the bottom

    def open_link(self, url: QUrl):
        """
        Open the clicked link in the default application.
        """
        self.log_console.append(f"DEBUG: Clicked URL: {url.toString()}")  # Debugging log
        if url.isLocalFile():
            QDesktopServices.openUrl(url)  # Open the local file in the default application
        else:
            self.log_message(f"Invalid link: {url.toString()}")  # Log invalid links for debugging

    def browse_excel_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Select Excel File", "", "Excel files (*.xlsx *.xlsm)")
        if file_path:
            self.excel_file_edit.setText(file_path)
            self.save_cached_value("excel_file", file_path)
            self.log_message(f"Selected Excel file: {file_path}")
            self.load_sheets(file_path)

    def load_sheets(self, file_path):
        """
        Populate the worksheet combo box based on the selected Excel file.
        """
        try:
            workbook = load_workbook(file_path, data_only=True)
            self.sheet_combo.clear()
            self.sheet_combo.addItems(workbook.sheetnames)
            self.log_message(f"Worksheets loaded: {', '.join(workbook.sheetnames)}",debug=True)
        except Exception as e:
            self.sheet_combo.clear()
            QMessageBox.critical(self, "Error", f"Failed to load worksheets: {e}")
            self.log_message(f"Error loading worksheets: {e}")

    def populate_headers(self):
        """
        Populate the Layer Name, Source Path, and Filter Category field combo boxes based on the selected worksheet and header row.
        """
        file_path = self.excel_file_edit.text()
        sheet_name = self.sheet_combo.currentText()
        header_row = self.header_row_spin.value()
        max_columns = self.max_columns_spin.value()

        if not file_path or not sheet_name:
            return

        try:
            workbook = load_workbook(file_path, data_only=True, read_only=True)
            sheet = workbook[sheet_name]

            # Read the header row
            headers = [
                cell.value for cell in sheet[header_row][:max_columns]
                if cell.value is not None
            ]

            self.layer_name_combo.clear()
            self.layer_name_combo.addItems(headers)

            self.source_path_combo.clear()
            self.source_path_combo.addItems(headers)

            self.filter_category_combo1.clear()
            self.filter_category_combo1.addItems(headers)

            self.filter_category_combo2.clear()
            self.filter_category_combo2.addItems(headers)

            self.log_message(f"Headers populated from row {header_row}: {', '.join(headers)}",debug=True)
        except Exception as e:
            self.layer_name_combo.clear()
            self.source_path_combo.clear()
            self.filter_category_combo1.clear()
            self.filter_category_combo2.clear()
            QMessageBox.critical(self, "Error", f"Failed to populate headers: {e}")
            self.log_message(f"Error populating headers: {e}")

    def populate_filter_categories1(self):
        """
        Populate the first filter category selection box based on the selected filter category field.
        """
        file_path = self.excel_file_edit.text()
        sheet_name = self.sheet_combo.currentText()
        header_row = self.header_row_spin.value()
        filter_category_field1 = self.filter_category_combo1.currentText()

        if not file_path or not sheet_name or not filter_category_field1:
            # Clear the filter category layout if no valid field is selected
            self.clear_filter_category_layout1()
            return

        try:
            workbook = load_workbook(file_path, data_only=True, read_only=True)
            sheet = workbook[sheet_name]

            # Get the column index for the filter category field
            headers = {cell.value: idx for idx, cell in enumerate(sheet[header_row], start=0)}
            if filter_category_field1 not in headers:
                raise ValueError("Selected filter category field not found in the header row.")

            filter_category_col1 = headers[filter_category_field1]

            # Extract unique values from the filter category column
            categories = set()
            for row in sheet.iter_rows(min_row=header_row + 1, max_row=sheet.max_row):
                value = row[filter_category_col1].value
                if value:
                    categories.add(str(value))
                if len(categories) >= 20:  # Limit to 20 categories
                    break

            # Clear the filter category layout before repopulating
            self.clear_filter_category_layout1()

            # Populate the scrollable filter category selection box
            for category in sorted(categories):
                checkbox = QCheckBox(category)
                self.filter_category_layout1.addWidget(checkbox)

            self.log_message(f"Filter categories populated: {', '.join(categories)}",debug=True)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to populate filter categories: {e}")
            self.log_message(f"Error populating filter categories: {e}")

    def populate_filter_categories2(self):
        """
        Populate the second filter category selection box based on the selected second filter category field.
        """
        file_path = self.excel_file_edit.text()
        sheet_name = self.sheet_combo.currentText()
        header_row = self.header_row_spin.value()
        filter_category_field2 = self.filter_category_combo2.currentText()

        if not file_path or not sheet_name or not filter_category_field2:
            # Clear the second filter category layout if no valid field is selected
            self.clear_filter_category_layout2()
            return

        try:
            workbook = load_workbook(file_path, data_only=True, read_only=True)
            sheet = workbook[sheet_name]

            # Get the column index for the second filter category field
            headers = {cell.value: idx for idx, cell in enumerate(sheet[header_row], start=0)}
            if filter_category_field2 not in headers:
                raise ValueError("Selected second filter category field not found in the header row.")

            filter_category_col2 = headers[filter_category_field2]

            # Extract unique values from the second filter category column
            categories = set()
            for row in sheet.iter_rows(min_row=header_row + 1, max_row=sheet.max_row):
                value = row[filter_category_col2].value
                if value:
                    categories.add(str(value))
                if len(categories) >= 20:  # Limit to 20 categories
                    break

            # Clear the second filter category layout before repopulating
            self.clear_filter_category_layout2()

            # Populate the scrollable second filter category selection box
            for category in sorted(categories):
                checkbox = QCheckBox(category)
                self.filter_category_layout2.addWidget(checkbox)

            self.log_message(f"Second filter categories populated: {', '.join(categories)}",debug=True)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to populate second filter categories: {e}")
            self.log_message(f"Error populating second filter categories: {e}")

    def clear_filter_category_layout1(self):
        """
        Clear all widgets from the first filter category layout and force a UI update.
        """
        self.log_message("Clearing filter category layout...", debug=True)
        for i in reversed(range(self.filter_category_layout1.count())):
            item = self.filter_category_layout1.itemAt(i)
            widget = item.widget()
            if widget:
                self.log_message(f"Removing widget: {widget.text()}", debug=True)
                self.filter_category_layout1.removeWidget(widget)
                widget.deleteLater()
        self.filter_category_layout1.update()  # Force the layout to update
        self.filter_category_widget1.update()  # Force the parent widget to update
        self.log_message("Filter category layout cleared.", debug=True)

    def clear_filter_category_layout2(self):
        """
        Clear all widgets from the second filter category layout and force a UI update.
        """
        self.log_message("Clearing second filter category layout...", debug=True)
        for i in reversed(range(self.filter_category_layout2.count())):
            item = self.filter_category_layout2.itemAt(i)
            widget = item.widget()
            if widget:
                self.log_message(f"Removing widget: {widget.text()}", debug=True)
                self.filter_category_layout2.removeWidget(widget)
                widget.deleteLater()
        self.filter_category_layout2.update()  # Force the layout to update
        self.filter_category_widget2.update()  # Force the parent widget to update
        self.log_message("Second filter category layout cleared.", debug=True)

    def browse_report_path(self):
        """
        Open a file dialog to select the validation report path.
        """
        file_path, _ = QFileDialog.getSaveFileName(self, "Select Validation Report Path", "", "CSV files (*.csv);;All files (*.*)")
        if file_path:
            self.report_path_edit.setText(file_path)
            self.save_cached_value("report_path", file_path)
            self.log_message(f"Selected validation report path: {file_path}",debug=True)

    def normalize_path(self, path, case_sensitive):
        """
        Normalize a file path or data source string for consistent comparison.
        """
        if not path:
            return ""

        # Detect if the path is URL-based (e.g., starts with "file:")
        if path.lower().startswith("file:"):
            # Remove the "file:" prefix
            normalized = path[5:]
            # remove / in front if still present
            while normalized.startswith("/"):
                normalized = normalized[1:]
            # Decode URL-encoded characters (e.g., %20 to space)
            normalized = unquote(normalized)
            # Remove everything after the ? symbol
            if "?" in normalized:
                normalized = normalized.split("?", 1)[0]
        else:
            # For non-URL paths, apply standard normalization
            normalized = path.strip().replace("\\", "/")

        # Remove everything after the | symbol
        if "|" in normalized:
            normalized = normalized.split("|", 1)[0]

        # Maintain case sensitivity if required
        if not case_sensitive:
            normalized = normalized.lower()

        # Remove redundant slashes
        while "//" in normalized:
            normalized = normalized.replace("//", "/")

        return normalized

    def validate_project(self):
        """
        Perform the validation logic and write the validation report.
        """
        # Save user selections to project variables
        self.save_cached_value("header_row", self.header_row_spin.value())
        self.save_cached_value("max_columns", self.max_columns_spin.value())
        self.save_cached_value("duplicate_match_mode", self.duplicate_match_combo.currentText())
        self.save_cached_value("sheet", self.sheet_combo.currentText())
        self.save_cached_value("layer_name_field", self.layer_name_combo.currentText())
        self.save_cached_value("source_path_field", self.source_path_combo.currentText())
        self.save_cached_value("use_filter_category1", str(self.use_filter_category_checkbox1.isChecked()))
        self.save_cached_value("filter_category_field1", self.filter_category_combo1.currentText())
        selected_categories1 = [
            checkbox.text() for i in range(self.filter_category_layout1.count())
            if isinstance((checkbox := self.filter_category_layout1.itemAt(i).widget()), QCheckBox) and checkbox.isChecked()
        ]
        self.save_cached_value("filter_categories1", ",".join(selected_categories1))

        self.save_cached_value("use_filter_category2", str(self.use_filter_category_checkbox2.isChecked()))
        self.save_cached_value("filter_category_field2", self.filter_category_combo2.currentText())
        selected_categories2 = [
            checkbox.text() for i in range(self.filter_category_layout2.count())
            if isinstance((checkbox := self.filter_category_layout2.itemAt(i).widget()), QCheckBox) and checkbox.isChecked()
        ]
        self.save_cached_value("filter_categories2", ",".join(selected_categories2))

        self.save_cached_value("layer_name_delimiter", self.layer_name_delimiter_edit.text())
        self.save_cached_value("case_sensitive_matching", str(self.case_sensitive_checkbox.isChecked()))
        self.save_cached_value("generate_html_report", str(self.generate_html_checkbox.isChecked()))
        self.save_cached_value("verbose_console", str(self.verbose_console_checkbox.isChecked()))
        layer_name_delimiter = self.layer_name_delimiter_edit.text()
        case_sensitive_matching = self.case_sensitive_checkbox.isChecked()

        excel_file = self.excel_file_edit.text()
        sheet_name = self.sheet_combo.currentText()
        header_row = self.header_row_spin.value()
        layer_name_field = self.layer_name_combo.currentText()
        source_path_field = self.source_path_combo.currentText()
        filter_category_field1 = self.filter_category_combo1.currentText()
        report_path = self.report_path_edit.text()
        duplicate_match_mode = self.duplicate_match_combo.currentText()
        use_filter1_category = self.use_filter_category_checkbox1.isChecked()
        use_filter2_category = self.use_filter_category_checkbox2.isChecked()
        filter_category_field2 = self.filter_category_combo2.currentText()

        if not excel_file or not sheet_name or not layer_name_field or not source_path_field or not report_path:
            QMessageBox.warning(self, "Missing Input", "Please fill in all fields before proceeding.")
            self.log_message("Validation aborted: Missing input fields.")
            return

        try:
            workbook = load_workbook(excel_file, data_only=True, read_only=True)
            sheet = workbook[sheet_name]

            # Map header names to column indices
            headers = {cell.value: idx for idx, cell in enumerate(sheet[header_row], start=0)}
            if layer_name_field not in headers or source_path_field not in headers or (use_filter1_category and filter_category_field1 not in headers) or (use_filter2_category and filter_category_field2 not in headers):
                raise ValueError("Selected fields not found in the header row.")

            layer_name_col = headers[layer_name_field]
            source_path_col = headers[source_path_field]
            filter_category_col1 = headers[filter_category_field1] if use_filter1_category else None
            filter_category_col2 = headers[filter_category_field2] if use_filter2_category else None

            # Extract all possible layers from the Excel file based on filters
            possible_layers = {}
            for row in sheet.iter_rows(min_row=header_row + 1, max_row=sheet.max_row):
                layer_name = row[layer_name_col].value
                source_path = row[source_path_col].value
                category1 = row[filter_category_col1].value if use_filter1_category else None
                category2 = row[filter_category_col2].value if use_filter2_category else None

                # Apply filters
                if use_filter1_category and (not category1 or str(category1) not in selected_categories1):
                    continue
                if use_filter2_category and (not category2 or str(category2) not in selected_categories2):
                    continue

                if layer_name and source_path:
                    normalized_layer_name = self.normalize_path(layer_name, case_sensitive_matching)
                    normalized_source_path = self.normalize_path(source_path, case_sensitive_matching)
                    possible_layers[normalized_layer_name] = {
                        "original_layer_name": layer_name,
                        "source_path": normalized_source_path,
                        "category1": category1,
                        "category2": category2,
                    }

            self.log_message(f"Possible layers extracted: {possible_layers}", debug=True)

            # Initialize unmatched layers with all possible layers
            unmatched_layers = possible_layers.copy()

            # Validate against project layers
            project_layers = QgsProject.instance().mapLayers()
            html_rows = []  # Collect rows for the HTML report
            matched_count = 0
            wrong_source_count = 0
            layer_name_not_found_count = 0
            total_count = 0

            for layer_id, layer in project_layers.items():
                layer_name = layer.name()
                layer_source = layer.dataProvider().dataSourceUri()

                # Normalize the layer name and source for comparison
                normalized_layer_name = self.normalize_path(layer_name, case_sensitive_matching)
                if layer_name_delimiter in normalized_layer_name:
                    normalized_layer_name = normalized_layer_name.split(layer_name_delimiter, 1)[0]
                    self.log_message(f"Layer name '{layer_name}' normalized to '{normalized_layer_name}' using delimiter '{layer_name_delimiter}'", debug=True)
                normalized_layer_source = self.normalize_path(layer_source, case_sensitive_matching)

                if normalized_layer_name in possible_layers:
                    reference_source = possible_layers[normalized_layer_name]["source_path"]
                    category1 = possible_layers[normalized_layer_name]["category1"]
                    category2 = possible_layers[normalized_layer_name]["category2"]
                    if normalized_layer_source == reference_source:
                        check_result = "MATCHED"
                        matched_count += 1
                    else:
                        check_result = "WRONGSOURCE"
                        wrong_source_count += 1

                    # Remove matched or wrong source layers from unmatched layers
                    unmatched_layers.pop(normalized_layer_name, None)
                else:
                    reference_source = ""
                    category1 = ""
                    category2 = ""
                    check_result = "LAYERNAMENOTFOUND"
                    layer_name_not_found_count += 1

                # Skip blank lines
                if not layer_name or not check_result or not normalized_layer_source:
                    self.log_message(f"Skipping blank line for layer '{layer_name}'", debug=True)
                    continue

                # Increment total count
                total_count += 1

                # Write the validation result to the report
                html_rows.append(
                    (layer_name, check_result, normalized_layer_source, reference_source, category1, category2)
                )

                self.log_message(f"Layer '{layer_name}' checked: {check_result}", debug=True)

            self.log_message(f"Unmatched layers after validation: {unmatched_layers}", debug=True)

            # Collect validation details
            from datetime import datetime
            import getpass
            report_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            report_user = getpass.getuser()
            qgis_project_path = QgsProject.instance().fileName()
            validation_excel_date_modified = datetime.fromtimestamp(os.path.getmtime(excel_file)).strftime("%Y-%m-%d %H:%M:%S")
            validation_filter_categories1 = ";".join(selected_categories1)
            validation_filter_categories2 = ";".join(selected_categories2)

            # Open the report file for writing (CSV)
            with open(report_path, "w", encoding="utf-8") as report_file:
                # Write the #VALIDATIONDETAILS# section
                report_file.write("#VALIDATIONDETAILS#\n")
                report_file.write(f"ReportDate={report_date}\n")
                report_file.write(f"ReportUser={report_user}\n")
                report_file.write(f"QGISprojectPath={qgis_project_path}\n")
                report_file.write(f"ValidationExcelFilePath={excel_file}\n")
                report_file.write(f"ValidationExcelSheet={sheet_name}\n")
                report_file.write(f"ValidationExcelDateModified={validation_excel_date_modified}\n")
                report_file.write(f"ValidationLayerNameField={layer_name_field}\n")
                report_file.write(f"ValidationSourceField={source_path_field}\n")
                report_file.write(f"ValidationFilterCatergories1={validation_filter_categories1}\n")
                report_file.write(f"ValidationFilterCatergories2={validation_filter_categories2}\n\n")

                # Write the #DATASOURCECHECK# section header
                report_file.write("#DATASOURCECHECK#\n")
                report_file.write("LayerName,CheckResult,LayerSource,ReferenceSource,FilterCategory,SecondFilterCategory\n")
                for row in html_rows:
                    report_file.write(",".join(map(str, row)) + "\n")

                # Write the #MISSINGVALIDATION# section
                report_file.write("\n#MISSINGVALIDATION#\n")
                report_file.write("LayerName,SourcePath,FilterCategory,SecondFilterCategory\n")
                for layer_data in unmatched_layers.values():
                    report_file.write(
                        f"{layer_data['original_layer_name']},{layer_data['source_path']},{layer_data['category1']},{layer_data['category2']}\n"
                    )

                # Write the #VALIDATIONSUMMARY# section
                report_file.write("\n#VALIDATIONSUMMARY#\n")
                report_file.write(f"DATASOURCES_MATCHED={matched_count}\n")
                report_file.write(f"DATASOURCES_WRONGSOURCE={wrong_source_count}\n")
                report_file.write(f"DATASOURCES_LAYERNAMENOTFOUND={layer_name_not_found_count}\n")
                report_file.write(f"DATASOURCES_TOTAL={total_count}\n")

            # Generate HTML report if the option is selected
            if self.generate_html_checkbox.isChecked():
                html_report_path = report_path.replace(".csv", ".html")
                with open(html_report_path, "w", encoding="utf-8") as html_file:
                    html_file.write("<html><head><title>Validation Report</title>")
                    html_file.write("<style>")
                    html_file.write("table { border-collapse: collapse; width: 100%; }")
                    html_file.write("th, td { border: 1px solid black; padding: 8px; text-align: left; }")
                    html_file.write("th { background-color: #f2f2f2; cursor: pointer; }")  # Add cursor pointer for sortable headers
                    html_file.write(".matched { background-color: #d4edda; }")  # Light green
                    html_file.write(".layernotfound { background-color: #fff3cd; }")  # Light yellow
                    html_file.write(".wrongsource { background-color: #f8d7da; }")  # Light red
                    html_file.write("</style>")
                    html_file.write("<script>")
                    html_file.write("""
                        // JavaScript function to sort table columns
                        function sortTable(tableId, columnIndex) {
                            const table = document.getElementById(tableId);
                            const rows = Array.from(table.rows).slice(1); // Exclude header row
                            const isAscending = table.getAttribute('data-sort-order') !== 'asc';
                            rows.sort((rowA, rowB) => {
                                const cellA = rowA.cells[columnIndex].innerText.toLowerCase();
                                const cellB = rowB.cells[columnIndex].innerText.toLowerCase();
                                if (cellA < cellB) return isAscending ? -1 : 1;
                                if (cellA > cellB) return isAscending ? 1 : -1;
                                return 0;
                            });
                            rows.forEach(row => table.tBodies[0].appendChild(row));
                            table.setAttribute('data-sort-order', isAscending ? 'asc' : 'desc');
                        }
                    """)
                    html_file.write("</script>")
                    html_file.write("</head><body>")
                    html_file.write("<h1>Validation Report</h1>")
                    html_file.write("<h2>Validation Details</h2>")
                    html_file.write(f"<p><b>Report Date:</b> {report_date}</p>")
                    html_file.write(f"<p><b>Report User:</b> {report_user}</p>")
                    html_file.write(f"<p><b>QGIS Project Path:</b> {qgis_project_path}</p>")
                    html_file.write(f"<p><b>Validation Excel File Path:</b> {excel_file}</p>")
                    html_file.write(f"<p><b>Validation Excel Sheet:</b> {sheet_name}</p>")
                    html_file.write(f"<p><b>Validation Excel Date Modified:</b> {validation_excel_date_modified}</p>")
                    html_file.write(f"<p><b>Validation Layer Name Field:</b> {layer_name_field}</p>")
                    html_file.write(f"<p><b>Validation Source Field:</b> {source_path_field}</p>")
                    html_file.write(f"<p><b>Validation Filter Categories 1:</b> {validation_filter_categories1}</p>")
                    html_file.write(f"<p><b>Validation Filter Categories 2:</b> {validation_filter_categories2}</p>")
                    html_file.write("<h2>Data Source Check</h2>")
                    html_file.write('<table id="datasource-check"><thead><tr>')
                    html_file.write('<th onclick="sortTable(\'datasource-check\', 0)">Layer Name</th>')
                    html_file.write('<th onclick="sortTable(\'datasource-check\', 1)">Check Result</th>')
                    html_file.write('<th onclick="sortTable(\'datasource-check\', 2)">Layer Source</th>')
                    html_file.write('<th onclick="sortTable(\'datasource-check\', 3)">Reference Source</th>')
                    html_file.write('<th onclick="sortTable(\'datasource-check\', 4)">Filter Category</th>')
                    html_file.write('<th onclick="sortTable(\'datasource-check\', 5)">Second Filter Category</th>')
                    html_file.write("</tr></thead><tbody>")
                    for layer_name, check_result, normalized_layer_source, reference_source, category1, category2 in html_rows:
                        result_class = ""
                        if check_result == "MATCHED":
                            result_class = "matched"
                        elif check_result == "LAYERNAMENOTFOUND":
                            result_class = "layernotfound"
                        elif check_result == "WRONGSOURCE":
                            result_class = "wrongsource"

                        html_file.write(
                            f"<tr><td>{layer_name}</td>"
                            f"<td class='{result_class}'>{check_result}</td>"
                            f"<td>{normalized_layer_source}</td>"
                            f"<td>{reference_source}</td>"
                            f"<td>{category1}</td>"
                            f"<td>{category2}</td></tr>"
                        )
                    html_file.write("</tbody></table>")

                    # Add the #MISSINGVALIDATION# section to the HTML report
                    html_file.write("<h2>Missing Validation Layers</h2>")
                    html_file.write('<table id="missing-validation"><thead><tr>')
                    html_file.write('<th onclick="sortTable(\'missing-validation\', 0)">Layer Name</th>')
                    html_file.write('<th onclick="sortTable(\'missing-validation\', 1)">Source Path</th>')
                    html_file.write('<th onclick="sortTable(\'missing-validation\', 2)">Filter Category</th>')
                    html_file.write('<th onclick="sortTable(\'missing-validation\', 3)">Second Filter Category</th>')
                    html_file.write("</tr></thead><tbody>")
                    for layer_data in unmatched_layers.values():
                        html_file.write(
                            f"<tr><td>{layer_data['original_layer_name']}</td>"
                            f"<td>{layer_data['source_path']}</td>"
                            f"<td>{layer_data['category1']}</td>"
                            f"<td>{layer_data['category2']}</td></tr>"
                        )
                    html_file.write("</tbody></table>")

                    html_file.write("<h2>Validation Summary</h2>")
                    html_file.write(f"<p><b>Matched:</b> <span style='color: #155724; background-color: #d4edda;'>Matched</span> ({matched_count})</p>")  # Light green
                    html_file.write(f"<p><b>Wrong Source:</b> <span style='color: #721c24; background-color: #f8d7da;'>Wrong Source</span> ({wrong_source_count})</p>")  # Light red
                    html_file.write(f"<p><b>Layer Name Not Found:</b> <span style='color: #856404; background-color: #fff3cd;'>Layer Name Not Found</span> ({layer_name_not_found_count})</p>")  # Light yellow
                    html_file.write(f"<p><b>Total:</b> {total_count}</p>")
                    html_file.write("</body></html>")

                self.log_message_link("HTML report written to:", html_report_path)

            QMessageBox.information(self, "Validation Complete", "Validation report generated successfully.")
        except Exception as e:
            # Capture the traceback details
            tb = traceback.format_exc()
            QMessageBox.critical(self, "Error", f"Validation failed: {e}\n\nDetails:\n{tb}")
            self.log_message(f"Validation failed: {e}\nTraceback:\n{tb}", debug=True)
